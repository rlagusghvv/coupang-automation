import json
import sys
from pathlib import Path

if len(sys.argv) < 3:
    print("usage: build_order_excel.py <input_json> <output_xlsx|output_xls>", file=sys.stderr)
    sys.exit(1)

in_path = Path(sys.argv[1])
out_path = Path(sys.argv[2])

payload = json.loads(in_path.read_text())
headers = payload.get("headers", [])
rows = payload.get("rows", [])

out_path.parent.mkdir(parents=True, exist_ok=True)

ext = out_path.suffix.lower()

if ext == ".xlsx":
    try:
        import openpyxl
    except Exception as e:
        print(f"openpyxl missing: {e}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(out_path)
    sys.exit(0)

if ext == ".xls":
    try:
        import xlwt
    except Exception as e:
        print(f"xlwt missing: {e}", file=sys.stderr)
        sys.exit(2)

    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")

    for col, name in enumerate(headers):
        ws.write(0, col, name)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row):
            ws.write(r_idx, c_idx, value)

    wb.save(str(out_path))
    sys.exit(0)

print(f"unsupported extension: {ext}", file=sys.stderr)
sys.exit(2)
